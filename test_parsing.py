from os import PathLike

from openpyxl import *
from tablib import *

from builtins import open


def openpyxl_test(table_address: str, worksheet: str) -> None:
    wb = load_workbook(table_address)
    ws = wb[worksheet]

    max_column = ws.max_column
    max_row = ws.max_row

    for row in range(1, (max_row + 1)):
        content = []
        for col in range(1, (max_column + 1)):
            if ws.cell(row, col).hyperlink is None:
                content.append(ws.cell(row, col).value)
            else:
                content.append(ws.cell(row, col).hyperlink.target)
        for _ in content:
            print(_, end='||')
        print()
        print()


def tablib_test(table_address: str | PathLike[str]) -> list[dict[str, str]]:
    """
    Проверка работы библиотеки tablib после изменения переменной row_vals в
    tablib.formats._xlsx.XLSXFormat.import_sheet со стандартного значения:
    row_vals = [c.value for c in row] на
    row_vals = [c.value if c.hyperlink is None else c.hyperlink.target for c in row]
    для чтения URL скрытого за текстом.
    """
    imported_data = Dataset()
    with open(table_address, 'rb') as fh:
        imported_data.load(fh, 'xlsx')
    return imported_data.dict


def tablib_dataset_xlsx(table_address: str | PathLike[str]) -> Dataset:
    imported_data = Dataset()
    with open(table_address, 'rb') as fh:
        dataset = imported_data.load(fh, 'xlsx')
    return dataset


def sales_row_coll_converter(dataset: Dataset,
                             number_of_columns: int = 6
                             ) -> Dataset:
    arr_data = dataset.dict
    new_arr_data = []
    months = (
        'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль',
        'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
    )
    years = ('2023', '2024')
    row = 0
    while row < len(arr_data):
        inter_dict = {}
        for heading in arr_data[row]:
            inter_dict[heading] = arr_data[row][heading]
            for year in years:
                for month in months:
                    if month in heading and year in heading:
                        inter_dict['Год'] = year
                        inter_dict['Месяц'] = month
                        inter_dict['Статус'] = inter_dict.pop(heading)

            if len(inter_dict) == number_of_columns:
                new_arr_data.append(inter_dict.copy())
        inter_dict.clear()
        row += 1
    new_dataset = Dataset()
    new_dataset.dict = new_arr_data
    return new_dataset


if __name__ == '__main__':
    # openpyxl_test(
    #     '/home/yolter/Downloads/тест/Занятость города 2023 тест.xlsx',
    #     'Статус')

    dataset = tablib_dataset_xlsx(
        '/home/yolter/Downloads/тест/билборды тест.xlsx')
    data = tablib_test_sales(dataset, 6)
    for _ in range(len(data)):
        print(data[_])

    # data = tablib_test('/home/yolter/Downloads/тест/стороны тест.xlsx')
    # for _ in range(len(data)):
    #     print(data[_])
    #     print(data)
